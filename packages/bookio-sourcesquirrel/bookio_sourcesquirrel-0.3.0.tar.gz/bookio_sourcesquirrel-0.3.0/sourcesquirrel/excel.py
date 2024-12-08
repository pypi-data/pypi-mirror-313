from typing import Iterable, Dict, List
from uuid import UUID

from openpyxl import load_workbook

from sourcesquirrel.classes.collection import Collection
from sourcesquirrel.classes.media_type import MediaType
from sourcesquirrel.classes.release_type import ReleaseType
from sourcesquirrel.classes.serie import Serie
from sourcesquirrel.constants.blockchains import ALGORAND, CARDANO, COINBASE, ETHEREUM, POLYGON
from sourcesquirrel.prefabs import blockchains, platforms
from sourcesquirrel.validators import algorand, cardano, evm

COLLECTIONS_SHEET = "collections"
SERIES_SHEET = "series"
MEDIA_TYPES_SHEET = "media types"
RELEASE_TYPES_SHEET = "release types"

ONCHAIN_COLLECTION_ID_VALIDATORS = {
    ALGORAND: algorand.is_address,
    CARDANO: cardano.is_policy_id,
    COINBASE: evm.is_address,
    ETHEREUM: evm.is_address,
    POLYGON: evm.is_address,
}


def read_rows(file_path: str, sheet_name: str, skip_header: bool = True):
    workbook = load_workbook(file_path, data_only=True)

    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' does not exist in the workbook.")

    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)

    if skip_header:
        next(rows)

    for row in rows:
        if row[0] is None:
            break

        yield row


def parse_list(value: str) -> List[str]:
    return list(filter(bool, map(str.strip, (value or "").split(","))))


def load_series(path: str) -> Iterable[Serie]:
    for row in read_rows(path, SERIES_SHEET):
        id = row[0]
        slug = row[1]
        name = row[2]
        emoji = row[3]

        blockchain = blockchains.get(row[4])

        serie = Serie(
            id=UUID(id),
            name=name,
            slug=slug,
            emoji=emoji,
            blockchain=blockchain,
        )

        serie.verify()
        yield serie


def load_media_types(path: str) -> Iterable[MediaType]:
    for row in read_rows(path, MEDIA_TYPES_SHEET):
        id = row[0]
        name = row[1]
        emoji = row[2]
        platform = platforms.get(row[3])

        media_type = MediaType(
            id=UUID(id),
            name=name,
            emoji=emoji,
            platform=platform,
        )

        media_type.verify()
        yield media_type


def load_release_types(path: str) -> Iterable[ReleaseType]:
    for row in read_rows(path, RELEASE_TYPES_SHEET):
        id = row[0]
        name = row[1]
        emoji = row[2]

        release_type = ReleaseType(
            id=UUID(id),
            name=name,
            emoji=emoji,
        )

        release_type.verify()
        yield release_type


def load_collections(
    path: str, all_series: Dict[str, Serie], all_media_types: Dict[str, MediaType], all_release_types: Dict[str, ReleaseType]
) -> Iterable[Collection]:
    for row in read_rows(path, COLLECTIONS_SHEET):
        id = row[0]
        blockchain = blockchains.get(row[1])
        media_type = all_media_types.get(row[2])
        onchain_collection_id = row[3]

        if "/" in onchain_collection_id:
            onchain_collection_id, onchain_collection_id_mainnet = onchain_collection_id.split("/")
        else:
            onchain_collection_id_mainnet = onchain_collection_id

        onchain_discriminator = row[4]
        title = row[5]
        is_public_domain = row[6]
        release_date = row[7].date()
        supply = int(row[8])
        cover_count = int(row[9])
        one_to_one_count = int(row[10])
        team_royalties = float(row[11])
        source_royalties = float(row[12])
        release_type = all_release_types.get(row[13])

        authors = parse_list(row[14])
        publishers = parse_list(row[15])
        series = [all_series[slug] for slug in parse_list(row[16])]

        collection = Collection(
            id=UUID(id),
            blockchain=blockchain,
            media_type=media_type,
            onchain_collection_id=onchain_collection_id,
            onchain_collection_id_mainnet=onchain_collection_id_mainnet,
            onchain_discriminator=onchain_discriminator,
            title=title,
            is_public_domain=is_public_domain,
            release_date=release_date,
            supply=supply,
            cover_count=cover_count,
            one_to_one_count=one_to_one_count,
            team_royalties=team_royalties,
            source_royalties=source_royalties,
            release_type=release_type,
            authors=authors,
            publishers=publishers,
            series=series,
        )

        collection.verify()
        yield collection
