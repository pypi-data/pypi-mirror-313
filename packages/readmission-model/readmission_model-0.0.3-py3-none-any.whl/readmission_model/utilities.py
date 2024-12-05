from typing import Literal
import yaml
import pathlib
import io
import os
import pandas as pd
import numpy as np
from scipy.sparse import csr_matrix
import msoffcrypto
import xlrd
import openpyxl
from readmission_model.model_utilities import *
from statistics import mean

HERE = pathlib.Path('..').resolve()

"""
Rate Functions
"""

def overall_readmission_rates_risk(df : pd.DataFrame, years : list):
    """
    Function to find the overall real readmission rates, the overall predicted readmission rates, and the overall O/E ratios for a list of years
    This function finds the predicted rates by summing the 'Risk' column

    Args:
        df (pd.Dataframe): The data to find the readmission rates from. This dataframe is the result of the 'Model' class found in Scripts.model_utilities.
                           The dataframe must have a column called 'Trigger_Readmission to hospital within 30 days' which contains the real readmissions.
                           The dataframe must have a column called 'Risk' which contains the probability that a admission be lead to a readmission.
                           The dataframe must have a column called 'AdmissionDate' which contains time values.
                           A dataframe like this is the result of train the model within the 'Model' class and evaluating it on data.
        years (list): A list of the years to find the rates for, listed as integers

    Returns:
        real_rates_overall (float): The real readmission rates for the whole data range
        pred_rates_risk_overall (float): The predicted readmission rates for the whole data range
        overall_o_e_risk (float): The O/E ratio for the whole data range

    """

    # Only consider the input date range 
    df = df[df['AdmissionDate'].dt.year.isin(years)]

    # Find true readmissions and probability for readmission
    y_true = df['Trigger_Readmission to hospital within 30 days']
    y_pred_prob = df['Risk']

    # Find predicted rate
    pred_rates_risk_overall = mean(y_pred_prob)

    # Find real rate
    real_rates_overall = mean(y_true)

    # Overall O/E ratio
    overall_o_e_risk = real_rates_overall/pred_rates_risk_overall

    return real_rates_overall, pred_rates_risk_overall, overall_o_e_risk


def overall_readmission_rates_threshold(df : pd.DataFrame, years : list):
    """
    Function to find the overall real readmission rates, the overall predicted readmission rates, and the overall O/E ratios for a list of years
    This function finds the predicted rates by finding a threshold and classifying based on that
    
    Args:
        df (pd.Dataframe): The data to find the readmission rates from. This dataframe is the result of the 'Model' class found in Scripts.model_utilities.
                           The dataframe must have a column called 'Trigger_Readmission to hospital within 30 days' which contains the real readmissions.
                           The dataframe must have a column called 'Risk' which contains the probability that a admission be lead to a readmission.
                           The dataframe must have a column called 'AdmissionDate' which contains time values.
                           A dataframe like this is the result of train the model within the 'Model' class and evaluating it on data.
        years (list): A list of the years to find the rates for, listed as integers
    
    Returns:
        real_rates_overall (float): The real readmission rates for the whole data range
        pred_rate_th_overall (float): The predicted readmission rates for the whole data range
        overall_o_e_th (float): The O/E ratio for the whole data range

    """

    # Only consider the input date range 
    df = df[df['AdmissionDate'].dt.year.isin(years)]

    # Find true readmissions and probability for readmission
    y_true = df['Trigger_Readmission to hospital within 30 days']
    y_pred_prob = df['Risk']

    # Find threshold to maximise the F score
    th, roc, roc_weighted, gmean, f = find_cutoffs(y_true, y_pred_prob, mean(y_true))

    # Create predictions based on that threshold
    y_pred_overall = [1 if i >= th else 0 for i in y_pred_prob] 

    # Find predicted rate
    pred_rate_th_overall = mean(y_pred_overall)

    # Find real rate
    real_rates_overall = mean(y_true)

    # Overall O/E ratio
    overall_o_e_th = real_rates_overall/pred_rate_th_overall

    return real_rates_overall, pred_rate_th_overall, overall_o_e_th


def find_readmission_rates_and_OE_risk(df : pd.DataFrame, years : list):
    """
    Function to find the real readmission rates, the predicted readmission rates, and the O/E ratios for a list of years

    Args:
        df (pd.Dataframe): The data to find the readmission rates from. This dataframe is the result of the 'Model' class found in Scripts.model_utilities.
                           The dataframe must have a column called 'Trigger_Readmission to hospital within 30 days' which contains the real readmissions.
                           The dataframe must have a column called 'Risk' which contains the probability that a admission be lead to a readmission.
                           The dataframe must have a column called 'AdmissionDate' which contains time values.
                           A dataframe like this is the result of train the model within the 'Model' class and evaluating it on data.
        years (list): A list of the years to find the rates for, listed as integers

    Returns:
        years (list): A list of the years to find the rates for, listed as strings
        real_rates (list): A list of the real readmission rates for each year, contains floats
        pred_rates (list): A list of the predicted readmission rates for each year, contains floats
        O_E_ratio (list): A list of the O/E ratio for each year, contains floats

    """

    df_by_year = {}
    true_and_risk = {}
    rates = {}
    rates_test = {}

    # Split data into different dataframes based on the year of their admission
    for year in years:
        df_by_year[year] = df[df['AdmissionDate'].dt.year == year]

    # Find rates for each year
    for year in years:
        rates[year] = {'y_true_rate' : mean(df_by_year[year]['Trigger_Readmission to hospital within 30 days']),
                       'y_pred_rate' : mean(df_by_year[year]['Risk']),
                       'O_E_ratio' : mean(df_by_year[year]['Trigger_Readmission to hospital within 30 days'])/mean(df_by_year[year]['Risk'])}
        
    # Convert the year values and the rates into lists
    years = list(rates.keys())
    real_rates = [rates[year]['y_true_rate'] for year in years]
    pred_rates = [rates[year]['y_pred_rate'] for year in years]
    O_E_ratio = [rates[year]['O_E_ratio'] for year in years]


    # For string conversion
    years = [str(year) for year in years]

    return years, real_rates, pred_rates, O_E_ratio


def find_readmission_rates_and_OE_threshold(df : pd.DataFrame, years : list):
    """
    Function to find the real readmission rates, the predicted readmission rates, and the O/E ratios for a list of years

    Args:
        df (pd.Dataframe): The data to find the readmission rates from. This dataframe is the result of the 'Model' class found in Scripts.model_utilities.
                           The dataframe must have a column called 'Trigger_Readmission to hospital within 30 days' which contains the real readmissions.
                           The dataframe must have a column called 'Risk' which contains the probability that a admission be lead to a readmission.
                           The dataframe must have a column called 'AdmissionDate' which contains time values.
                           A dataframe like this is the result of train the model within the 'Model' class and evaluating it on data.
        years (list): A list of the years to find the rates for, listed as integers

    Returns:
        years (list): A list of the years to find the rates for, listed as strings
        real_rates (list): A list of the real readmission rates for each year, contains floats
        pred_rates (list): A list of the predicted readmission rates for each year, contains floats
        O_E_ratio (list): A list of the O/E ratio for each year, contains floats

    """

    df_by_year = {}
    true_and_prob = {}
    true_and_pred = {}
    rates = {}

    # Split data into different dataframes based on the year of their admission
    for year in years:
        df_by_year[year] = df[df['AdmissionDate'].dt.year == year]

    # Find true readmissions and prediction probability for each year
    for year in df_by_year.keys():
        true_and_prob[year] = {'y_true':df_by_year[year]['Trigger_Readmission to hospital within 30 days'],
                               'y_pred_prob':df_by_year[year]['Risk']}

    # Find the threshold for each year and predict readmission
    for year in true_and_prob.keys():
        threshold, roc, roc_weighted, gmean, f0_5 = find_cutoffs(true_and_prob[year]['y_true'], 
                                                                 true_and_prob[year]['y_pred_prob'], 
                                                                 true_and_prob[year]['y_true'].mean())
        true_and_pred[year] = {'y_true':list(true_and_prob[year]['y_true']),
                               'y_pred':[1 if i >= threshold else 0 for i in true_and_prob[year]['y_pred_prob']]}
        
    
    # Find the real readmission rate and the predicted admission rate for each year, along with the O/E ratio
    for year in true_and_pred.keys():
        rates[year] = {'RealRate':mean(true_and_pred[year]['y_true']),
                       'PredRate':mean(true_and_pred[year]['y_pred']),
                       'O_E_ratio':mean(true_and_pred[year]['y_true'])/mean(true_and_pred[year]['y_pred'])}
        
    # Convert the year values and the rates into lists
    years = list(rates.keys())
    real_rates = [rates[year]['RealRate'] for year in years]
    pred_rates = [rates[year]['PredRate'] for year in years]
    O_E_ratio = [rates[year]['O_E_ratio'] for year in years]

    # For string conversion
    years = [str(year) for year in years]

    return years, real_rates, pred_rates, O_E_ratio

def find_readmission_rates_and_OE_constant_threshold(df : pd.DataFrame, years : list):
    """
    Function to find the real readmission rates, the predicted readmission rates, and the O/E ratios for a list of years

    Args:
        df (pd.Dataframe): The data to find the readmission rates from. This dataframe is the result of the 'Model' class found in Scripts.model_utilities.
                           The dataframe must have a column called 'Trigger_Readmission to hospital within 30 days' which contains the real readmissions.
                           The dataframe must have a column called 'Risk' which contains the probability that a admission be lead to a readmission.
                           The dataframe must have a column called 'AdmissionDate' which contains time values.
                           A dataframe like this is the result of train the model within the 'Model' class and evaluating it on data.
        years (list): A list of the years to find the rates for, listed as integers

    Returns:
        years (list): A list of the years to find the rates for, listed as strings
        real_rates (list): A list of the real readmission rates for each year, contains floats
        pred_rates (list): A list of the predicted readmission rates for each year, contains floats
        O_E_ratio (list): A list of the O/E ratio for each year, contains floats

    """

    # Only consider the input date range 
    df = df[df['AdmissionDate'].dt.year.isin(years)]

    # Find true readmissions and probability for readmission
    y_true = df['Trigger_Readmission to hospital within 30 days']
    y_pred_prob = df['Risk']

    # Find threshold to maximise the F score
    th, roc, roc_weighted, gmean, f = find_cutoffs(y_true, y_pred_prob, mean(y_true))

    df_by_year = {}
    true_and_prob = {}
    true_and_pred = {}
    rates = {}

    # Split data into different dataframes based on the year of their admission
    for year in years:
        df_by_year[year] = df[df['AdmissionDate'].dt.year == year]

    # Find true readmissions and prediction probability for each year
    for year in df_by_year.keys():
        true_and_prob[year] = {'y_true':df_by_year[year]['Trigger_Readmission to hospital within 30 days'],
                               'y_pred_prob':df_by_year[year]['Risk']}

    # Find the threshold for each year and predict readmission
    for year in true_and_prob.keys():
        true_and_pred[year] = {'y_true':list(true_and_prob[year]['y_true']),
                               'y_pred':[1 if i >= th else 0 for i in true_and_prob[year]['y_pred_prob']]}
        
    
    # Find the real readmission rate and the predicted admission rate for each year, along with the O/E ratio
    for year in true_and_pred.keys():
        rates[year] = {'RealRate':mean(true_and_pred[year]['y_true']),
                       'PredRate':mean(true_and_pred[year]['y_pred']),
                       'O_E_ratio':mean(true_and_pred[year]['y_true'])/mean(true_and_pred[year]['y_pred'])}
        
    # Convert the year values and the rates into lists
    years = list(rates.keys())
    real_rates = [rates[year]['RealRate'] for year in years]
    pred_rates = [rates[year]['PredRate'] for year in years]
    O_E_ratio = [rates[year]['O_E_ratio'] for year in years]

    # For string conversion
    years = [str(year) for year in years]

    return years, real_rates, pred_rates, O_E_ratio


def readmission_dataframe(years : list, real_rate : list, pred_rate : list, O_E_ratio : list, 
                          overall_real_rate : float, overall_pred_rate : float, overall_O_E_ratio : float):

    """
    Function to take the readmission rates and O/E ratios by year as well as from the overall time period and
    create a dataframe of all the results, including a emphasized row

    Args:
        years (list): A list of the years where the rates have been found, listed as integers 
        real_rate (list): A list of real readmission rates by year, listed as floats
        pred_rate (list): A list of predicted readmission rates by year, listed as floats
        O_E_ratio (list): A list of O/E ratios by year, listed as floats
        overall_real_rate (float): The real readmission rate from the whole time range
        overall_pred_rate (float): The predicted readmission rate from the whole time range
        overall_O_E_ratio (float): The O/E ratio from the whole time range

    Returns:
        df (pd.DataFrame): A dataframe with all values included by year, with an 'overall' row added highlighted in grey
    """
    # Create df for the overall data
    overall_row = pd.DataFrame({'Year' : 'Overall',
                                'Real Readmission Rate' : overall_real_rate,
                                'Predicted Readmission Rate' : overall_pred_rate,
                                'O/E Ratio' : overall_O_E_ratio}, index = [0])
    
    # Create df for year by year data
    rates_df = pd.DataFrame({'Year' : pd.Series(years), 
                             'Real Readmission Rate' : pd.Series(real_rate),
                             'Predicted Readmission Rate' : pd.Series(pred_rate),
                             'O/E Ratio' : pd.Series(O_E_ratio)})
    
    # Join the two dfs together
    rates_df_overall = pd.concat([rates_df, overall_row]).reset_index(drop=True)

    # Define the 'overall' row  to emphasize
    row_to_emphasize = rates_df_overall.shape[0]-1  # index of the row to highlight

    # Create a function to apply the style
    def highlight_row(s):
        return ['background-color: lightsteelblue' if i == row_to_emphasize else '' for i in range(len(s))]

    # Apply the styling
    df = rates_df_overall.style.apply(highlight_row, axis=0)

    return df

"""
READING / WRITING FILE FUNCTIONS
"""

class PTLDataFrame(pd.DataFrame):
    """A class for dataframes that are specific to PTL project."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

    @classmethod
    def from_encrypted_spreadsheet(
        cls,
        path,
        pw_file: str | pathlib.Path | None = None,
        sheet_name: int | str = 0,
        **kwargs,
    ) -> "PTLDataFrame":
        """Create a PTLDataFrame from an encrypted excel file."""

        if not str(path).endswith(".xlsx"):
            raise Exception("File must be an excel file")

        path = pathlib.Path(path)
        filename = path.name

        if not pw_file:
            pw_file = pathlib.Path(path).parent.parent.parent / "sql_pw_demo.yaml"

        try:
            pw = yaml.safe_load(open(pw_file))[filename]
        except KeyError:
            pw = None
        except FileNotFoundError:
            raise Exception(f"Password file not found at {pw_file}")

        # Assume that if no password found, spreadsheet is not encrypted
        if not pw:
            try:
                df = pd.read_excel(path, sheet_name=sheet_name, **kwargs)
            except ValueError:
                raise Exception("No password found and spreadsheet not readable")
            except xlrd.XLRDError:
                raise Exception("No password found and spreadsheet not readable")
            return cls(df)

        # print(f"Using password file {pw_file}")
        decrypted_workbook = io.BytesIO()
        with open(path, "rb") as file:
            office_file = msoffcrypto.OfficeFile(file)
            office_file.load_key(password=pw)
            office_file.decrypt(decrypted_workbook)

        df = pd.read_excel(
            decrypted_workbook,
            engine="openpyxl",
            sheet_name=sheet_name,
            **kwargs,
        )
        return cls(df)
    
    
def concatenate_excel_sheets(workbook_path):
    """
    Concatenate multiple sheets with same columns into single dataframe.
    Args:
        workbook_path : excel file location
    """
    wb = openpyxl.load_workbook(workbook_path)
    dfs = []
    for sheet in wb.sheetnames:
        if sheet.lower() == "version control":
            continue
        df_temp = PTLDataFrame.from_encrypted_spreadsheet(
            workbook_path, sheet_name=sheet
        )
        df_temp["Sheet"] = sheet
        dfs.append(df_temp)
    df = pd.concat(dfs)
    return df


def read_all_sheets(workbook_path):
    """
    Iterate through all sheets in excel file and return dataframes in dictionary.
    Args:
        workbook_path : excel file location
    """
    wb = openpyxl.load_workbook(workbook_path)
    dfs = {}
    for sheet in wb.sheetnames:
        if sheet.lower() == "version control":
            continue
        dfs[sheet] = PTLDataFrame.from_encrypted_spreadsheet(
            workbook_path, sheet_name=sheet
        )
    return dfs


def save_as_sparse(df : pd.DataFrame, save_path, fill_value=0):
    """
    Save dataframe as sparse matrix where possible and remaining as dense matrix at save_path.
    Args:
        df (pd.DataFrame) : dataframe to save
        save_path : path to save location
        fill_value : the value to omit. Defaults to 0.
    """
    df_sparse = df.astype(pd.SparseDtype("float", fill_value))
    df_sparse.to_csv(save_path, index=False)


def calc_if_not_exists(
    path: str | pathlib.Path,
    dataframe: pd.DataFrame,
    function,
    **kwargs: dict,
) -> pd.DataFrame:
    """_summary_

    Args:
        path (str | pathlib.Path): _description_
        dataframe (pd.DataFrame): _description_
        function (_type_): _description_

    Returns:
        _type_: _description_
    """
    if os.path.exists(path):
        return pd.read_csv(path)

    calced_df = function(dataframe, **kwargs)
    calced_df.to_csv(path)
    return calced_df



# OPCS / ICD10 FUNCTIONS
    
#ICD10 = PTLDataFrame.from_encrypted_spreadsheet(HERE / "files/CRABCodes.xlsx",sheet_name="Diagnostics")


def get_opcs_cat(proc: str, exclude_XYZ: bool = False) -> float | str:
    """_summary_

    Args:
        proc (str): _description_
        exclude_XYZ (bool, optional): _description_. Defaults to False.

    Returns:
        float | str: _description_
    """
    if type(proc) == float:
        return np.nan

    first_letter = proc[0]
    match first_letter:
        case "A":
            return "Nervous System"
        case "B":
            return "Endocrine System and Breast"
        case "C":
            return "Eye"
        case "D":
            return "Ear"
        case "E":
            return "Respiratory Tract"
        case "F":
            return "Mouth"
        case "G":
            return "Upper Digestive System"
        case "H":
            return "Lower Digestive System"
        case "J":
            return "Other Abdominal Organs, Principally Digestive"
        case "K":
            return "Heart"
        case "L":
            return "Arteries and Veins"
        case "M":
            return "Urinary"
        case "N":
            return "Male Genital Organs"
        case "P":
            return "Lower Female Genital Tract"
        case "Q":
            return "Upper Female Genital Tract"
        case "R":
            return "Female Genital Tract Associated with Pregnancy, Childbirth and the Puerperium"
        case "S":
            return "Skin"
        case "T":
            return "Soft Tissue"
        case "U":
            return "Diagnostic Imaging, Testing and Rehabilitation"
        case "V":
            return "Bones and Joints of Skull and Spine"
        case "W":
            return "Other Bones and Joints"
        case "X" if exclude_XYZ == False:
            return "Miscellaneous Operations"
        case "Y" if exclude_XYZ == False:
            return "Subsidiary Classification of Methods of Operation"
        case "Z" if exclude_XYZ == False:
            return "Subsidiary Classification of Sites of Operation"
        case _:
            return "None of the above - procedures"


def get_icd_cat(diag: str) -> float | str:
    """_summary_

    Args:
        diag (str): _description_

    Returns:
        float | str: _description_
    """
    if type(diag) == float:
        return np.nan

    first_letter = diag[0]
    match first_letter:
        case "A", "B":
            return "Certain infectious and parasitic diseases"
        case "C", "D" if int(diag[1:3]) < 50:
            return "Neoplasms"
        case "D" if int(diag[1:3]) >= 50:
            return "Diseases of the blood and blood-forming organs and certain disorders involving the immune mechanism"
        case "E":
            return "Endocrine, nutritional and metabolic diseases"
        case "F":
            return "Mental, Behavioural and Neurodevelopmental disorders"
        case "G":
            return "Diseases of the nervous system"
        case "H" if int(diag[1:3]) < 60:
            return "Diseases of the eye and adnexa"
        case "H" if int(diag[1:3]) >= 60:
            return "Diseases of the ear and mastoid process"
        case "I":
            return "Diseases of the circulatory system"
        case "J":
            return "Diseases of the respiratory system"
        case "K":
            return "Diseases of the digestive system"
        case "L":
            return "Diseases of the skin and subcutaneous tissue"
        case "M":
            return "Diseases of the musculoskeletal system and connective tissue"
        case "N":
            return "Diseases of the genitourinary system"
        case "O":
            return "Pregnancy, childbirth and the puerperium"
        case "P":
            return "Certain conditions originating in the perinatal period"
        case "Q":
            return (
                "Congenital malformations, deformations and chromosomal abnormalities"
            )
        case "R":
            return "Symptoms, signs and abnormal clinical and laboratory findings, not elsewhere classified"
        case "S", "T":
            return "Injury, poisoning and certain other consequences of external causes"
        case "V", "W", "X", "Y":
            return "External causes of morbidity and mortality"
        case "Z":
            return "Factors influencing health status and contact with health services"
        case _:
            return "None of the above - diagnoses"


# PROCESSING CODE MAPPING FUNCTIONS

def squash_codes(
    df: pd.DataFrame,
    code: Literal["Proc"] | Literal["Diag"],
) -> pd.DataFrame:
    """_summary_

    Args:
        df (pd.DataFrame): _description_
        code (Literal[&quot;Proc&quot;] | Literal[&quot;Diag&quot;]): _description_

    Returns:
        pd.DataFrame: _description_
    """
    code_columns = df.filter(like=code, axis=1)

    def concat_columns(row) -> str:
        tmp_set = set(
            row[col]
            for col in code_columns.columns
            if not "nan" in str(row[col]).lower()
        )
        tmp_list = list(tmp_set)
        tmp_list.sort()
        return "_".join(tmp_list)

    df[code + "Codes"] = df.apply(concat_columns, axis=1)
    return df


def pivot_cols_to_rows(
    df: pd.DataFrame,
    cols_to_pivot: list,
    init_cols: list = ["OperationID"],
    new_col: str = "Code",
) -> pd.DataFrame:
    """
    Pivot columns to rows.
    Args:
        df (pd.DataFrame): Dataframe with columns to pivot.
        cols (list): Columns to be pivoted to rows.
        init_cols (list, optional): List of initial columns to keep as identifiers in each row. Defaults to ["OperationID"].
        new_col (str, optional): Name of new column, Defaults to "Code".
    Returns:
        pd.DataFrame: Pivoted dataframe with identifier columns, and a column containing row by row values of the pivoted columns.
    """
    df_new = pd.DataFrame(columns=init_cols + [new_col])
    for col in cols_to_pivot:
        df_new_tmp = df.loc[~(df[col].isna())][init_cols + [col]].copy()
        df_new_tmp.rename(columns={col: new_col}, inplace=True)
        df_new = pd.concat([df_new, df_new_tmp], ignore_index=True)
    # Remove null rows, remove duplicate codes per operation (and append descriptions from procedure and diagnostic code list)
    df_new = df_new.loc[~(df_new[new_col].isna())].drop_duplicates()
    return df_new


def get_codes(
    df: pd.DataFrame,
    code: Literal["diag"] | Literal["proc"],
    init_cols: list =["OperationID"],
    code_dataframe: None | pd.DataFrame = None,
    # expand_X = False
) -> pd.DataFrame:
    """_summary_

    Args:
        df (pd.DataFrame): _description_
        code (Literal[&quot;Diag&quot;] | Literal[&quot;Proc&quot;]): _description_
        init_cols (list, optional): _description_. Defaults to ["OperationID"].
        code_dataframe (None | pd.DataFrame, optional): _description_. Defaults to None.
        expand_X (bool): whether to add all 4 character codes with same base 3 when 4 char codes end in 'X', e.g. I10X also adds I101,I102...

    Returns:
        pd.DataFrame: _description_
    """

    """
    # add 4 char codes for X codes
    if expand_X==True:
        df_sub = df_codes.loc[df_codes["Code"].str[-1]=='X']
        df_sub.loc[:,"Code"] = df_sub["Code"].str[:3]
        for i in np.arange(10):
            df_add = pd.DataFrame(np.repeat(df_sub.values, 10, axis=0))
            df_add.columns = df_sub.columns
            df_add["Code"] = df_add["Code"].apply(lambda x: str(x)+str(i))
            df_codes = pd.concat([df_codes,df_add],ignore_index=True)
    """
    # pivot columns to rows
    cols_to_pivot = [i for i in df.columns if (code in i.lower()) == True and (len(i) <= 6)]
    df_codes = pivot_cols_to_rows(df,cols_to_pivot,init_cols=init_cols,new_col="Code")
    
    # append opcs/icd10 descriptions
    if code_dataframe is not None:
        df_codes = pd.merge(df_codes, code_dataframe, on="Code", how="left")
        
    return df_codes


def fill_3char_codes(df, code_col):
    """Input dataframe containing 3 character codes or 4 character code with * or % as fourth character,
    to return dataframe with filled in rows with 4 character codes ending in 0-9 and X
    """

    # TODO check this is the desired behaviour
    def check_length_is_3_or_star(x):
        if type(x) == float:
            return False
        return len(x) == 3 or "*" in str(x) or "%" in str(x)

    stars = pd.DataFrame()
    for i in list(np.arange(10)) + ["X"]:
        star = df.loc[df[code_col].apply(check_length_is_3_or_star)].copy()

        star[code_col] = star[code_col].apply(lambda x: x[:3] + str(i))
        stars = pd.concat([stars, star], ignore_index=True)
    df = pd.concat([df, stars], ignore_index=True)
    df = df.loc[
        (df[code_col].str.contains("\*") == False)
        & (df[code_col].str.contains("%") == False)
    ]

    return df


def pivot_list_to_rows(df, col_to_pivot, col_new_name="Code"):
    """
    Pivots a dataframe from a single column of a list of codes to a dataframe with a row per code.
    Inputs: dataframe, column name to pivot, new column name after pivoting
    """
    # TODO - use explode
    df_keep = df.loc[(df[col_to_pivot].isna())]
    df_pivot = df.loc[~(df[col_to_pivot].isna())]
    n = df_pivot.columns.shape[0]

    # Create a column for each code
    df_expand = pd.merge(
        df_pivot,
        df_pivot[col_to_pivot].str.split(",", expand=True),
        left_index=True,
        right_index=True,
    )
    n_expand = df_expand.columns.shape[0]
    n_cols = n_expand - n
    df_expand.columns = list(df_expand.columns[:n].values) + [
        "Code" + str(i) for i in range(n_cols)
    ]
    code_cols = [i for i in df_expand.columns if "Code" in i]
    for col in code_cols:
        df_expand[col] = df_expand[col].apply(lambda x: str(x).replace(" ", ""))

    # Pivot multiple columns to rows
    df_all = pd.DataFrame()
    for i in range(n_cols):
        df_tmp = df_expand[df_expand.columns[:-n_cols]].copy()
        df_tmp[col_new_name] = df_expand["Code" + str(i)]
        df_all = pd.concat([df_all, df_tmp], ignore_index=True)
    df_all = df_all.loc[
        (df_all[col_new_name].isna() == False)
        & (df_all[col_new_name] != "None")
        & (df_all[col_new_name] != "nan")
    ]

    # Reappend to unpivoted df
    df_all = pd.concat([df_all, df_keep], ignore_index=True)
    return df_all


def add_X_codes(df: pd.DataFrame, icd10, code_col: str = "Code"):
    """
    Adds additional row for ...X codes, where it exists in ICD10 dataframe.
    Inputs: dataframe, ICD10 dataframe, column name with codes.
    """
    # get unit 3 char codes
    df_sub = df.copy()
    df_sub.loc[:, code_col] = df_sub[code_col].str[:3]
    df_sub.drop_duplicates(inplace=True)

    x_codes = icd10.loc[icd10["Code"].str[-1] == "X"]
    x_codes.loc[:, "Code"] = x_codes["Code"].str[:3]

    df_sub = df_sub.loc[df_sub[code_col].isin(x_codes["Code"])]
    df_sub.loc[:, code_col] = df_sub.loc[:, code_col].apply(lambda x: str(x) + "X")

    df = pd.concat([df, df_sub])
    return df



if __name__ == "__main__":
    df1 = PTLDataFrame.from_encrypted_spreadsheet(
        path="../files/CRABCodes.xlsx", sheet_name="Diagnostics"
    )

    df2 = PTLDataFrame.from_encrypted_spreadsheet(
        path="../files/CoMorbidity_List_ICD10_final_L2S2_June2022_v3.xlsx"
    )

    print(get_icd_cat("A00"))
    print(get_opcs_cat("A00"))
