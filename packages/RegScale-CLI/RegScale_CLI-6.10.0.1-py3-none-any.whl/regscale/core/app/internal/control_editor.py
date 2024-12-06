#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Module to allow user to make changes to Control Implementations in an Excel spreadsheet for a user-friendly experience
"""

import os
import shutil
import sys
from pathlib import Path
from typing import Union, Any

import click

# standard python imports
import math
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation

from regscale.core.app.api import Api
from regscale.core.app.application import Application
from regscale.core.app.logz import create_logger
from regscale.core.app.utils.app_utils import (
    check_empty_nan,
    check_file_path,
    error_and_exit,
    get_current_datetime,
    get_user_names,
)
from regscale.models.app_models.click import regscale_id, regscale_module
from regscale.models.regscale_models.control import Control
from regscale.models.regscale_models.control_implementation import ControlImplementation


@click.group(name="control_editor")
def control_editor():
    """
    Performs actions on Control Editor Feature to edit controls to RegScale.
    """


# Get data and pull into Excel worksheets.


@control_editor.command(name="generate")
@regscale_id()
@regscale_module()
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path for created excel files to be saved to.",
    default=Path("./artifacts"),
    required=True,
)
def generate_data_download(regscale_id: int, regscale_module: str, path: Path):
    """
    This function will build and populate a spreadsheet of all control implementations
    with the selected RegScale Parent Id and RegScale Module.
    """
    data_load(regscale_id=regscale_id, regscale_module=regscale_module, path=path)


def data_load(regscale_id: int, regscale_module: str, path: Path) -> None:
    """Function takes organizer record and module and build excel worksheet of control implementations.

    :param int regscale_id: RegScale Parent Id
    :param str regscale_module: RegScale Parent Module
    :param Path path: directory of file location
    :rtype: None
    """
    import pandas as pd  # Optimize import performance

    logger = create_logger()
    api = Api()

    # Making directory for files

    check_file_path(path)

    workbook = Workbook()
    ws = workbook.active
    ws.title = f"Impls_PId({regscale_id}_{regscale_module})"
    workbook.create_sheet("Accounts")

    workbook.save(filename=os.path.join(path, "all_implementations.xlsx"))
    shutil.copy(
        os.path.join(path, "all_implementations.xlsx"),
        os.path.join(path, "old_implementations.xlsx"),
    )

    # Loading data from RegScale database into two workbooks.

    body = """
            query{
                controlImplementations (skip: 0, take: 50, where: {parentId: {eq: parent_id} parentModule: {eq: "parent_module"}}) {
                    items {
                        id
                        controlID
                        controlOwner {
                            firstName
                            lastName
                            userName
                        }
                        control {
                            title
                            description
                            controlId
                        }
                        status
                        policy
                        implementation
                        responsibility
                        inheritable
                        parentId
                        parentModule
                    }
                    totalCount
                    pageInfo {
                        hasNextPage
                    }
                }
            }""".replace(
        "parent_module", regscale_module
    ).replace(
        "parent_id", str(regscale_id)
    )

    existing_implementation_data = api.graph(query=body)

    if existing_implementation_data["controlImplementations"]["totalCount"] > 0:
        raw_data = existing_implementation_data["controlImplementations"]["items"]

        all_imps = []
        for item in raw_data:
            Id = item["id"]
            ControlId = item["controlID"]
            ControlOwner = (
                str(item["controlOwner"]["lastName"]).strip()
                + ", "
                + str(item["controlOwner"]["firstName"]).strip()
                + " ("
                + str(item["controlOwner"]["userName"]).strip()
                + ")"
            )
            ControlName = item["control"]["controlId"]
            ControlTitle = item["control"]["title"]
            Description = item["control"]["description"]
            Status = item["status"]
            Policy = item["policy"]
            Implementation = item["implementation"]
            Responsibility = item["responsibility"]
            Inheritable = item["inheritable"]

            all_imps.append(
                [
                    Id,
                    ControlId,
                    ControlOwner,
                    ControlName,
                    ControlTitle,
                    Description,
                    Status,
                    Policy,
                    Implementation,
                    Responsibility,
                    Inheritable,
                ]
            )

        all_imps_df = pd.DataFrame(
            all_imps,
            columns=[
                "Id",
                "ControlId",
                "ControlOwner",
                "ControlName",
                "ControlTitle",
                "Description",
                "Status",
                "Policy",
                "Implementation",
                "Responsibility",
                "Inheritable",
            ],
        )

        with pd.ExcelWriter(
            os.path.join(path, "all_implementations.xlsx"),
            mode="a",
            engine="openpyxl",
            if_sheet_exists="overlay",
        ) as writer:
            all_imps_df.to_excel(
                writer,
                sheet_name=f"Impls_PId({regscale_id}_{regscale_module})",
                index=False,
            )
            get_user_names().to_excel(
                writer,
                sheet_name="Accounts",
                index=False,
            )

        with pd.ExcelWriter(
            os.path.join(path, "old_implementations.xlsx"),
            mode="a",
            engine="openpyxl",
            if_sheet_exists="overlay",
        ) as writer:
            all_imps_df.to_excel(
                writer,
                sheet_name=f"Impls_PId({regscale_id}_{regscale_module})",
                index=False,
            )
    else:
        error_and_exit("No records exist for the given RegScale Id and RegScale Module.")

    # Adding Data validation to "old_implementations.xlsx" file that will be used as reference.

    workbook2 = load_workbook(os.path.join(path, "old_implementations.xlsx"))
    worksheet2 = workbook2.active
    worksheet2.protection.sheet = True
    workbook2.save(filename=os.path.join(path, "old_implementations.xlsx"))

    # Adding Data Validation to "all_implementations.xlsx" file to be adjusted internally by clients.

    workbook = load_workbook(os.path.join(path, "all_implementations.xlsx"))
    worksheet = workbook.active
    worksheet.protection.sheet = True
    accounts_worksheet = workbook["Accounts"]
    accounts_worksheet.protection.sheet = True

    dv1 = DataValidation(
        type="list",
        formula1='"Not Implemented, Fully Implemented, In Remediation, Not Applicable, Inherited, Planned"',
        allow_blank=True,
        showDropDown=False,
        error="Your entry is not one of the available options",
        errorTitle="Invalid Entry",
        prompt="Please select from the list",
    )
    dv2 = DataValidation(
        type="list",
        formula1='"Provider, Customer, Shared, Not Applicable"',
        allow_blank=True,
        showDropDown=False,
        error="Your entry is not one of the available options",
        errorTitle="Invalid Entry",
        prompt="Please select from the list",
    )
    dv3 = DataValidation(type="list", formula1='"TRUE, FALSE"', allow_blank=True)
    dv4 = DataValidation(
        type="list",
        formula1="=Accounts!$A$2:$A$" + str(get_maximum_rows(sheet_object=workbook["Accounts"])),
        allow_blank=False,
        showDropDown=False,
        error="Your entry is not one of the available options",
        errorTitle="Invalid Entry",
        prompt="Please select from the list",
    )

    worksheet.add_data_validation(dv1)
    worksheet.add_data_validation(dv2)
    worksheet.add_data_validation(dv3)
    worksheet.add_data_validation(dv4)
    dv1.add("G2:G1048576")
    dv2.add("J2:J1048576")
    dv3.add("K2:K1048576")
    dv4.add("C2:C1048576")

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        highlight_fill = PatternFill(
            start_color="7C7C7C", end_color="7C7C7C", fill_type="solid"
        )  # Adding highlights to column that can be edited
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))

        adjusted_width = (max_length + 2) * 1.2
        if adjusted_width < 50:
            worksheet.column_dimensions[column].width = adjusted_width
        else:
            worksheet.column_dimensions[column].width = 50

        if column in [
            "E",
            "F",
            "I",
        ]:
            for cell in col:
                cell.alignment = Alignment(wrap_text=True)

        if column in [
            "C",
            "G",
            "H",
            "I",
            "J",
            "K",
        ]:  # Check if current column is column to edit
            for cell in col:
                cell.fill = highlight_fill  # Apply the highlight fill
                cell.protection = Protection(locked=False)  # Unprotect the cell

    workbook.save(filename=os.path.join(path, "all_implementations.xlsx"))

    logger.info("Successfully created the directory %s.", path)
    logger.info("All files are located within directory.")

    logger.info(
        "Your data has been loaded into your excel workbook. Please open the all_implementations workbook and make your desired changes."
    )
    return None


# Save Spreadsheet if file changed, append Update API changes that were manually entered in an Excel worksheet


@control_editor.command(name="load")
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path where excel workbooks are located.",
    default=Path("./artifacts"),
    required=True,
)
@click.option(
    "--skip_prompt",
    type=click.BOOL,
    help="To Skip (Y/N) Prompt, input True.",
    default=False,
    required=False,
)
def generate_db_update(path: Path, skip_prompt: bool):
    """
    This function will check changes made to spreadsheet and upload any changes made to RegScale.

    """
    db_update(path, skip_prompt)


def db_update(path: Path, skip_prompt: bool = True) -> None:
    """Function will check changes made by user and upload any changes to RegScale.

    :param Path path: directory of file location
    :param bool skip_prompt: boolean to skip prompt save message, defaults to True
    :rtype: None
    """
    import pandas as pd  # Optimize import performance
    import numpy as np  # Optimize import performance

    logger = create_logger()

    logger.info("Proceed only after you have made the necessary changes and have saved file.")

    x = "y" if skip_prompt else input("Ready to Proceed (Y/N): ").lower()

    if x[0] == "y":
        file_path = os.path.join(path, "all_implementations.xlsx")
        if not os.path.exists(file_path):
            error_and_exit(f"Unable to locate the file {file_path}.")

        df = load_workbook(file_path)

        sheet_name = df.sheetnames[0]
        sheet_name = sheet_name[sheet_name.find("(") + 1 : sheet_name.find(")")].split("_")
        # set the variables to the correct values
        for item in set(sheet_name):
            try:
                regscale_parent_id = int(item)
            except ValueError:
                regscale_module = item

        df1 = pd.read_excel(os.path.join(path, "all_implementations.xlsx"), sheet_name=0, index_col="Id")

        df2 = pd.read_excel(os.path.join(path, "old_implementations.xlsx"), sheet_name=0, index_col="Id")

        if df1.equals(df2):
            logger.warning("No differences detected.")
            sys.exit(0)

        else:
            logger.warning("*** WARNING *** Differences Found.")

            # Logs changes to txt file

            diff_mask = (df1 != df2) & ~(df1.isnull() & df2.isnull())
            ne_stacked = diff_mask.stack()
            changed = ne_stacked[ne_stacked]
            changed.index.names = ["Id", "Column"]
            difference_locations = np.where(diff_mask)
            changed_to = df1.values[difference_locations]
            changed_from = df2.values[difference_locations]
            changes = pd.DataFrame({"From": changed_from, "To": changed_to}, index=changed.index)
            changes.to_csv(
                os.path.join(path, "differences.txt"),
                header=True,
                index=True,
                sep=" ",
                mode="a",
            )

            upload_data(regscale_parent_id, regscale_module, path)

    logger.info("Please check differences.txt file located in artifacts folder to see changes made.")
    return None


def upload_data(regscale_parent_id: int, regscale_module: str, path: Path) -> None:
    """
    Batch uploads updated control implementation statements to the provided RegScale parent ID.

    :param int regscale_parent_id: RegScale parent ID
    :param str regscale_module: RegScale parent module
    :param Path path: file path where control spreadsheet resides
    :rtype: None
    """
    import pandas as pd  # Optimize import performance

    app = Application()
    config = app.config
    api = Api()

    diff = pd.read_csv(os.path.join(path, "differences.txt"), header=0, sep=" ", index_col=None)
    ids = []
    for i, row in diff.iterrows():
        ids.append(row["Id"])

    id_df = pd.DataFrame(ids, index=None, columns=["Id"])
    id_df2 = id_df.drop_duplicates()

    reader = pd.read_excel(os.path.join(path, "all_implementations.xlsx"))
    accounts = pd.read_excel(os.path.join(path, "all_implementations.xlsx"), sheet_name="Accounts")
    accounts = accounts.rename(columns={"User": "ControlOwner", "UserId": "ControlOwnerId"})

    updates = reader[reader["Id"].isin(id_df2["Id"])]
    updates = updates.merge(accounts, how="left", on="ControlOwner")
    updates = updates.T.to_dict()

    updated_implementations = [
        build_implementation(i, regscale_parent_id, regscale_module, app) for i in updates.values()
    ]

    api.update_server(
        url=config["domain"] + "/api/controlImplementation",
        json_list=updated_implementations,
        message="Working on uploading updated control implementations to RegScale.",
        config=config,
        method="put",
    )


def build_implementation(i: dict, regscale_parent_id: int, regscale_module: str, app: Application) -> dict:
    """
    Builds a ControlImplementation object from a dictionary

    :param dict i: dictionary of control implementation
    :param int regscale_parent_id: RegScale parent ID
    :param str regscale_module: RegScale parent module
    :param Application app: Application object
    :return: ControlImplementation object
    :rtype: dict
    """
    control = Control(
        title=i["ControlTitle"],
        description=i["Description"],
        controlId=i["ControlName"],
    )
    control_implementation = ControlImplementation(
        id=i["Id"],
        controlOwnerId=i["ControlOwnerId"],
        control=control.dict(),
        status=i["Status"],
        implementation=check_empty_nan(i["Implementation"]),
        policy=check_empty_nan(i["Policy"]),
        controlID=i["ControlId"],
        responsibility=check_empty_nan(i["Responsibility"]),
        parentId=regscale_parent_id,
        parentModule=regscale_module,
        inheritable=check_inheritable(i["Inheritable"]),
        lastUpdatedById=app.config["userId"],
        dateLastUpdated=get_current_datetime(),
    )
    return control_implementation.dict()


# Delete and remove files from user's system.
@control_editor.command(name="delete_files")
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path of file location.",
    default=Path("./artifacts"),
    required=True,
)
def generate_delete_file(path: Path):
    """This command will delete files used during the Control editing process."""
    delete_file(path)


def delete_file(path: Path) -> int:
    """
    Deletes files used during the process

    :param Path path: directory of file location
    :return: Number of files deleted
    :rtype: int
    """
    logger = create_logger()
    file_names = [
        "all_implementations.xlsx",
        "old_implementations.xlsx",
        "differences.txt",
    ]
    deleted_files = []

    for file_name in file_names:
        if os.path.isfile(path / file_name):
            os.remove(path / file_name)
            deleted_files.append(file_name)
        else:
            logger.warning("No %s file found. Checking for other files before exiting.", file_name)
    logger.info("%i files have been deleted: %s", len(deleted_files), ", ".join(deleted_files))
    return len(deleted_files)


def check_inheritable(
    value: Any,
) -> Union[
    str, float, bool
]:  # this function has to be checked separate to account for API only accepting False Boolean unlike other class params
    """This function takes a given value for an inheritable and checks if value is empty or NaN based on value type.

    :param Any value: A string or float object
    :return: A string value, float value or False
    :rtype: Union[str, float, bool]
    """
    if isinstance(value, str) and value.strip() == "":
        return False
    if isinstance(value, float) and math.isnan(value):
        return False
    return value


def get_maximum_rows(*, sheet_object: Any) -> int:
    """This function finds the last row containing data in a spreadsheet

    :param Any sheet_object: excel worksheet to be referenced
    :return: integer representing last row with data in spreadsheet
    :rtype: int
    """
    return sum(any(col.value is not None for col in row) for max_row, row in enumerate(sheet_object, 1))
