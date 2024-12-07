import logging
import random
import pandas as pd
import re
import importlib.resources
import os
import tempfile
from faker import Faker
from decimal import Decimal
import openpyxl


from nemo_library.features.config import Config
from nemo_library.features.fileingestion import ReUploadFile
from nemo_library.features.focus import focusMoveAttributeBefore
from nemo_library.features.projects import (
    LoadReport,
    createImportedColumn,
    createOrUpdateReport,
    createOrUpdateRule,
    createProject,
    getImportedColumns,
    getProjectList,
)
from nemo_library.utils.utils import display_name, import_name, internal_name, log_error

__all__ = ["updateProjectsForMigMan"]

CALCULATED_FIELD = "CALCULATED_FIELD"


def updateProjectsForMigMan(
    config: Config,
    projects: list[str] = None,
    proALPHA_project_status_file: str = None,
    csv_files_directory: str = None,
    multi_projects: dict[str, str] = None,
) -> None:
    """
    Creates projects for MigMan based on a provided project list or a proALPHA project status file.

    Args:
        config (Config): Configuration object containing connection details and headers.
        projects (list[str], optional): A list of project names to create. Defaults to None.
        proALPHA_project_status_file (str, optional): Path to a proALPHA migration status file
                                                      to generate the project list. Defaults to None.

    Returns:
        None

    Raises:
        ValueError: If no projects are provided or the project list is empty.
        RuntimeError: If no matching files are found for a given project.

    Notes:
        - If `proALPHA_project_status_file` is provided, the project list is generated from the file.
        - Each project is processed by:
            - Finding matching files.
            - Sorting the files.
            - Extracting a postfix and processing each file using `process_file`.
        - Logs errors and exceptions if files are missing or the project list is invalid.
    """

    if proALPHA_project_status_file:
        projects = getNEMOStepsFrompAMigrationStatusFile(proALPHA_project_status_file)

    if not projects or len(projects) == 0:
        log_error("no projects given")

    for project in projects:
        logging.info(f"Scanning project '{project}'")

        # Get matching files
        matching_files = get_matching_files(project)
        if not matching_files:
            log_error(f"No files found for project '{project}'. Please check for typos")

        # Sort files
        sorted_files = sort_files(matching_files)

        # Process each file
        for filename in sorted_files:
            postfix = extract_postfix(filename)
            process_file(
                config=config,
                filename=filename,
                postfix=postfix,
                csv_files_directory=csv_files_directory,
                multi_projects=multi_projects,
            )


def get_matching_files(project: str) -> list[str]:
    """Get all matching files for the given project."""
    pattern = re.compile(
        f"^Template {re.escape(project)} (MAIN|ADD\\d+)\\.csv$", re.IGNORECASE
    )
    return [
        resource
        for resource in importlib.resources.contents("nemo_library.migmantemplates")
        if pattern.match(resource)
    ]


def sort_files(files: list[str]) -> list[str]:
    """Sort files with 'MAIN' first, followed by 'ADD' files in numerical order."""
    return sorted(
        files,
        key=lambda x: (
            0 if "MAIN" in x else 1,
            int(re.search(r"ADD(\d+)", x).group(1)) if "ADD" in x else float("inf"),
        ),
    )


def extract_postfix(filename: str) -> str:
    """Extract the postfix from the filename."""
    match = re.search(r"(\w{4})\.", filename)
    if match:
        return match.group(1)
    else:
        error_message = f"No valid postfix found in filename '{filename}'"
        logging.error(error_message)
        raise ValueError(error_message)


def process_file(
    config: Config,
    filename: str,
    postfix: str,
    csv_files_directory: str = None,
    multi_projects: dict[str, str] = None,
) -> None:
    """Process a single file and create a project."""

    pattern = re.compile(
        r"^Template (?P<project>.*?) (MAIN|Add\d+)\.csv$", re.IGNORECASE
    )
    match = pattern.match(filename)
    if match:
        project = match.group("project")
    else:
        log_error(f"filename '{filename}' does not match expected pattern")

    logging.info(
        f"Processing file {filename} (postfix {postfix}) for project '{project}'"
    )
    with importlib.resources.open_binary(
        "nemo_library.migmantemplates", filename
    ) as file:

        # there is a bug in the migman exporter. It has too many separators and the columns are "shifted". This is the csv in plain text
        # Column;Column Name;Description / Remark;Location in proALPHA;Data Type;Format
        # 1;Kunde;Customer;S_Kunde.Kunde;INTEGER;zzzzzzzz;
        # 2;Adressnummer;Address Number;S_Adresse.AdressNr;INTEGER;zzzzzzz9;
        # 3;Name;Name 1;S_Adresse.Name1;CHARACTER;x(80);

        df = pd.read_csv(
            file,
            skiprows=2,
            encoding="ISO-8859-1",
            sep=";",
        )

        df["Format"] = df["Data Type"]
        df["Data Type"] = df["Location in proALPHA"]
        df["Location in proALPHA"] = df["Description / Remark"]
        df.drop(columns=["Description / Remark"], inplace=True)

        df.loc[df["Location in proALPHA"].isna(), "Column Name"] = df["Column"]
        df.loc[df["Location in proALPHA"].isna(), "Data Type"] = "CHARACTER"
        df.loc[df["Location in proALPHA"].isna(), "Format"] = "x(20)"
        df.loc[df["Location in proALPHA"].isna(), "Location in proALPHA"] = (
            CALCULATED_FIELD + " " + df.loc[df["Location in proALPHA"].isna(), "Column"]
        )

        addons = (
            multi_projects[project]
            if multi_projects and project in multi_projects.keys()
            else [""]
        )
        for addon in addons:
            projectname = (
                f"{project} {addon} {'(' + postfix + ')' if postfix != 'MAIN' else ''}"
            )
            projectname = re.sub(r"\s+", " ", projectname)
            projectname = projectname.strip()

            # project already exists?
            new_project = False
            if not projectname in getProjectList(config)["displayName"].to_list():
                new_project = True
                logging.info(f"Project not found in NEMO. Create it...")
                createProject(
                    config=config,
                    projectname=projectname,
                    description=f"Data Model for Mig Man table '{project}'",
                )

                process_columns(config, projectname, df)

                updateReports(config, projectname, df)

            uploadData(
                config, projectname, df, csv_files_directory, postfix, new_project
            )


def process_columns(
    config: Config,
    projectname: str,
    df: pd.DataFrame,
) -> None:

    # synchronize columns
    lastDisplayName = None
    for idx, row in df.iterrows():

        logging.info(f"process column {row["Column"]} with index {idx}")
        displayName = display_name(row["Location in proALPHA"], idx)
        internalName = internal_name(row["Location in proALPHA"], idx)
        importName = import_name(row["Location in proALPHA"], idx)

        # column already exists?
        if True:
            logging.info(
                f"Colunn {displayName} not found in project {projectname}. Create it."
            )
            description = "\n".join([f"{key}: {value}" for key, value in row.items()])
            createImportedColumn(
                config=config,
                projectname=projectname,
                displayName=displayName,
                internalName=internalName,
                importName=importName,
                dataType="string",  # we try a string-only-project to avoid importing errors
                description=description,
            )

            logging.info(f"Move column {displayName} to position {idx} in focus")
            focusMoveAttributeBefore(
                config=config,
                projectname=projectname,
                sourceDisplayName=displayName,
                targetDisplayName=lastDisplayName,
            )

        lastDisplayName = displayName


def uploadData(
    config: Config,
    projectname: str,
    df: pd.DataFrame,
    csv_files_directory: str = None,
    postfix: str = None,
    new_project: bool = False,
) -> None:

    # "real" data given? let's take this instead of the dummy file
    if csv_files_directory:
        if not postfix:
            postfix = "MAIN"
        if postfix == "MAIN":
            postfix = ""
        postfix = (" " + postfix).strip()
        file_path = os.path.join(csv_files_directory, f"{projectname}{postfix}.csv")
        if os.path.exists(file_path):
            uploadRealData(config, projectname, df, file_path)
            return

    # otherwise, we upload dummy data, but only for newly created projects
    if new_project:
        uploadDummyData(config, projectname, df)


def uploadRealData(
    config: Config,
    projectname: str,
    df: pd.DataFrame,
    file_path: str,
):
    logging.info(f"file '{file_path}' found - load real data from this file")

    # ensure that the CSV file has proper formatting. All columns must be defined, but not all are needed

    # Read the first line of the CSV file to get column names
    with open(file_path, "r") as file:
        first_line = file.readline().strip()

    csv_display_names = first_line.split(";")
    csv_display_names = [
        x.strip().strip('"').replace("\ufeff", "") for x in csv_display_names
    ]

    nemo_import_names = [
        import_name(row["Location in proALPHA"], idx) for idx, row in df.iterrows()
    ]

    for csv_display_name in csv_display_names:
        if not csv_display_name in nemo_import_names:
            log_error(f"column '{csv_display_name}' not defined.")

    # now, we are sure that the file is formattted property. Load file and add missing columns
    df = pd.read_csv(file_path, sep=";")

    # Add missing columns
    # Identify missing columns
    missing_columns = [col for col in nemo_import_names if col not in df.columns]

    if missing_columns:

        # Create a DataFrame with missing columns and default values
        missing_df = pd.DataFrame({col: [""] * len(df) for col in missing_columns})

        # Concatenate the original DataFrame with the missing columns
        df = pd.concat([df, missing_df], axis=1)

    # Write to a temporary file and upload
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_file_path = os.path.join(temp_dir, "tempfile.csv")

        df.to_csv(
            temp_file_path,
            index=False,
            sep=";",
            na_rep="",
        )
        logging.info(
            f"dummy file {temp_file_path} written for project '{projectname}'. Uploading data to NEMO now..."
        )

        ReUploadFile(
            config=config,
            projectname=projectname,
            filename=temp_file_path,
            update_project_settings=False,
        )
        logging.info(f"upload to project {projectname} completed")


def uploadDummyData(
    config: Config,
    projectname: str,
    df: pd.DataFrame,
):
    logging.info(f"Upload dummy file for project {projectname}")

    faker = Faker("de_DE")

    columns = [
        import_name(row["Location in proALPHA"], idx) for idx, row in df.iterrows()
    ]
    data_types = df["Data Type"].to_list()
    formats = df["Format"].to_list()

    def generate_column_data(data_type, format, num_rows):

        match data_type.lower():
            case "character":
                # Parse format to get maximum length
                match = re.search(r"x\((\d+)\)", format)
                field_length = int(match.group(1)) if match else len(format)

                # Generate fake texts now
                if field_length <= 6:
                    return [
                        faker.word()[:field_length].strip() for _ in range(num_rows)
                    ]
                else:
                    return [
                        faker.text()[:field_length].strip() for _ in range(num_rows)
                    ]

            case "integer":
                # Parse format
                negative_numbers = "-" in format
                if negative_numbers:
                    format = format.replace("-", "")

                match = re.search(r"z\((\d+)\)", format)
                field_length = int(match.group(1)) if match else len(format)

                def generate_integer(negative_numbers, field_length):
                    # Generate the maximum integer based on the field length
                    max_integer = 10**field_length - 1

                    # Generate a random integer within the allowed range
                    result = faker.random_int(min=0, max=max_integer)

                    # Randomly make the number negative if negative numbers are allowed
                    if negative_numbers and random.choice([True, False]):
                        result *= -1

                    return result

                return [
                    generate_integer(negative_numbers, field_length)
                    for _ in range(num_rows)
                ]
            case "decimal":

                # Parse format
                negative_numbers = "-" in format
                if negative_numbers:
                    format = format.replace("-", "")

                # decimals?
                decimals = len(format.split(".")[1]) if "." in format else 0
                if decimals > 0:
                    format = format[: len(format) - decimals - 1]

                # now the format has the length without decimals and without sign

                # Parse format to get maximum length
                match = re.search(r"z\((\d+)\)", format)
                field_length = int(match.group(1)) if match else len(format)

                def generate_decimal(negative_numbers, field_length, decimals):
                    # Generate the maximum integer based on the field length
                    max_integer = 10**field_length - 1
                    # Define the range for decimal places
                    max_decimal = 10**decimals
                    # Generate a random integer within the allowed range
                    integer_part = faker.random_int(min=0, max=max_integer)
                    # Generate random decimal places
                    decimal_part = faker.random_int(min=0, max=max_decimal - 1)

                    # Combine the integer and decimal parts into a Decimal object
                    result = Decimal(f"{integer_part}.{decimal_part:0{decimals}d}")

                    # Randomly make the number negative if negative numbers are allowed
                    if negative_numbers and random.choice([True, False]):
                        result *= -1

                    return result

                return [
                    generate_decimal(negative_numbers, field_length, decimals)
                    for _ in range(num_rows)
                ]
            case "date" | "datetime" | "datetime-tz":
                return [faker.date_this_decade() for _ in range(num_rows)]
            case "logical":
                format_options = format.split("/")
                return [
                    (
                        format_options[0]
                        if random.choice([True, False])
                        else format_options[1]
                    )
                    for _ in range(num_rows)
                ]
            case _:
                return [faker.text() for _ in range(num_rows)]

    # Introduce errors into the dataset
    def introduce_errors(df, error_rate=0.2):
        num_rows, num_cols = df.shape
        num_errors = int(num_rows * num_cols * error_rate)
        for _ in range(num_errors):
            row_idx = random.randint(0, num_rows - 1)
            col_idx = random.randint(0, num_cols - 1)
            col_name = df.columns[col_idx]
            dtype = data_types[col_idx]
            format = formats[col_idx]

            match dtype.lower():
                case "character":
                    match = re.search(r"x\((\d+)\)", format)
                    field_length = int(match.group(1)) if match else len(format)

                    if random.choice([True, False]):
                        df.at[row_idx, col_name] = ""  # Empty field
                    else:
                        df.at[row_idx, col_name] = "X" * (field_length + 5)

                case "integer" | "decimal":
                    negative_numbers = "-" in format

                    if not negative_numbers and random.choice([True, False]):
                        df.at[row_idx, col_name] = -99999  # Out-of-range value
                    else:

                        if random.choice([True, False]):
                            df.at[row_idx, col_name] = None  # Missing value
                        else:
                            df.at[row_idx, col_name] = (
                                "INVALID NUMBER"  # Invalid format
                            )

                case "date" | "datetime" | "datetime-tz":
                    if random.choice([True, False]):
                        df.at[row_idx, col_name] = None  # Missing value
                    else:
                        df.at[row_idx, col_name] = "INVALID_DATE"  # Invalid format
                case "logical":
                    if random.choice([True, False]):
                        df.at[row_idx, col_name] = None  # Missing value
                    else:
                        df.at[row_idx, col_name] = "INVALID_BOOLEAN"  # Invalid format
                case _:
                    df.at[row_idx, col_name] = ""  # Empty field

        return df

    num_rows = 500
    data = {
        column: generate_column_data(
            data_type=data_type, format=format, num_rows=num_rows
        )
        for column, format, data_type in zip(columns, formats, data_types)
    }

    dummy_df = pd.DataFrame(data)

    for idx, col in enumerate(columns):
        dummy_df[col] = dummy_df[col].astype("str")

    # Introduce 10% errors into the dummy dataset
    dummy_df = introduce_errors(dummy_df)

    # Write to a temporary file and upload
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_file_path = os.path.join(temp_dir, "tempfile.csv")

        dummy_df.to_csv(
            temp_file_path,
            index=False,
            sep=";",
            na_rep="",
        )
        logging.info(
            f"dummy file {temp_file_path} written for project '{projectname}'. Uploading data to NEMO now..."
        )

        ReUploadFile(
            config=config,
            projectname=projectname,
            filename=temp_file_path,
            update_project_settings=False,
        )
        logging.info(f"upload to project {projectname} completed")


def updateReports(
    config: Config,
    projectname: str,
    df: pd.DataFrame,
) -> None:

    logging.info(
        f"Update deficiency mining reports and rules for project {projectname}"
    )

    # create deficiency mining reports
    displayNames = [
        display_name(row["Location in proALPHA"], idx) for idx, row in df.iterrows()
    ]
    internalNames = [
        internal_name(row["Location in proALPHA"], idx) for idx, row in df.iterrows()
    ]
    dataTypes = df["Data Type"].to_list()
    formats = df["Format"].to_list()

    # create column specific fragments
    frags_checked = []
    frags_msg = []
    for displayName, columnNameInternal, dataType, format in zip(
        displayNames, internalNames, dataTypes, formats
    ):

        frag_check = []
        frag_msg = []

        # # mandatory fields
        # if row["pflicht"] == "ja":
        #     frag_check.append(
        #         f"{column_name_internal} IS NULL or TRIM({column_name_internal}) = ''"
        #     )
        #     frag_msg.append(f"{column_name} is mandatory")

        # data type specific checks
        match dataType.lower():
            case "character":
                # Parse format to get maximum length
                match = re.search(r"x\((\d+)\)", format)
                field_length = int(match.group(1)) if match else len(format)
                frag_check.append(f"LENGTH({columnNameInternal}) > {field_length}")
                frag_msg.append(
                    f"{displayName} exceeds field length (max {field_length} digits)"
                )

            case "integer" | "decimal":
                # Parse format
                negative_numbers = "-" in format
                if negative_numbers:
                    format = format.replace("-", "")

                if not negative_numbers:
                    frag_check.append(
                        f"LEFT(TRIM({columnNameInternal}), 1) = '-' OR RIGHT(TRIM({columnNameInternal}), 1) = '-'"
                    )
                    frag_msg.append(f"{displayName} must not be negative")

                # decimals?
                decimals = len(format.split(".")[1]) if "." in format else 0
                if decimals > 0:
                    format = format[: len(format) - decimals - 1]
                    frag_check.append(
                        f"""LOCATE(TO_VARCHAR(TRIM({columnNameInternal})), '.') > 0 AND 
            LENGTH(RIGHT(TO_VARCHAR(TRIM({columnNameInternal})), 
                        LENGTH(TO_VARCHAR(TRIM({columnNameInternal}))) - 
                        LOCATE(TO_VARCHAR(TRIM({columnNameInternal})), '.'))) > {decimals}"""
                    )
                    frag_msg.append(
                        f"{displayName} has too many decimals ({decimals} allowed)"
                    )

                match = re.search(r"z\((\d+)\)", format)
                field_length = int(match.group(1)) if match else len(format)

                frag_check.append(
                    f"""LENGTH(
                    LEFT(
                        REPLACE(TO_VARCHAR(TRIM({columnNameInternal})), '-', ''), 
                        LOCATE('.', REPLACE(TO_VARCHAR(TRIM({columnNameInternal})), '-', '')) - 1
                    )
                ) > {field_length}"""
                )
                frag_msg.append(
                    f"{displayName} has too many digits before the decimal point ({field_length} allowed)"
                )

                frag_check.append(
                    f"NOT {columnNameInternal} LIKE_REGEXPR('^[-]?[0-9]+(\\.[0-9]+)?[-]?$')"
                )
                frag_msg.append(f"{displayName} is not a valid number")

            case "date":
                frag_check.append(
                    f"NOT {columnNameInternal} LIKE_REGEXPR('^(\\d{{4}})-(0[1-9]|1[0-2])-(0[1-9]|[1-2][0-9]|3[0-1])$')"
                )
                frag_msg.append(f"{displayName} is not a valid date")

            case "logical":
                format_for_regex = format.replace("/", "|")
                frag_check.append(
                    f"NOT {columnNameInternal} LIKE_REGEXPR('^({format_for_regex})$')"
                )
                frag_msg.append(
                    f'{displayName}: logical value does not match format "{format}"'
                )

        # special fields

        if "mail" in columnNameInternal:
            frag_check.append(
                f"NOT {columnNameInternal} LIKE_REGEXPR('^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\\.[a-zA-Z]{{2,}}$')"
            )
            frag_msg.append(f"{displayName} is not a valid email")

        if (
            ("phone" in columnNameInternal)
            or ("telefon" in columnNameInternal)
            or ("fax" in columnNameInternal)
        ):
            frag_check.append(
                f"NOT {columnNameInternal} LIKE_REGEXPR('^\\+?[0-9\\s\\-()]{5,15}$')"
            )
            frag_msg.append(f"{displayName} is not a valid phone number")

        # now build deficiency mining report for this column (if there are checks)
        if frag_check:

            # save checks and messages for total report
            frags_checked.extend(frag_check)
            frags_msg.extend(frag_msg)
            sorted_columns = [columnNameInternal] + [
                col for col in internalNames if col != columnNameInternal
            ]

            # case statements for messages and dm report
            case_statement_specific = " ||\n\t".join(
                [
                    f"CASE\n\t\tWHEN {check}\n\t\tTHEN CHAR(10) || '{msg}'\n\t\tELSE ''\n\tEND"
                    for check, msg in zip(frag_check, frag_msg)
                ]
            )

            status_conditions = " OR ".join(frag_check)

            sql_statement = f"""SELECT
\t{',\n\t'.join(sorted_columns)}
\t,LTRIM({case_statement_specific},CHAR(10)) AS DEFICIENCY_MININNG_MESSAGE
\t,CASE 
\t\tWHEN {status_conditions} THEN 'check'
    ELSE 'ok'
END AS STATUS
FROM
$schema.$table"""

            # create the report
            report_display_name = f"(DEFICIENCIES) {displayName}"
            report_internal_name = internal_name(report_display_name)

            createOrUpdateReport(
                config=config,
                projectname=projectname,
                displayName=report_display_name,
                internalName=report_internal_name,
                querySyntax=sql_statement,
                description=f"Deficiency Mining Report for column '{displayName}' in project '{projectname}'",
            )

            createOrUpdateRule(
                config=config,
                projectname=projectname,
                displayName=displayName,
                ruleSourceInternalName=report_internal_name,
                ruleGroup="02 Columns",
                description=f"Deficiency Mining Rule for column '{displayName}' in project '{projectname}'",
            )

        logging.info(
            f"project: {projectname}, column: {displayName}: {len(frag_check)} frags added"
        )

    # now setup global dm report and rule

    # case statements for messages and dm report
    case_statement_specific = " ||\n\t".join(
        [
            f"CASE\n\t\tWHEN {check}\n\t\tTHEN  CHAR(10) || '{msg}'\n\t\tELSE ''\n\tEND"
            for check, msg in zip(frags_checked, frags_msg)
        ]
    )

    status_conditions = " OR ".join(frags_checked)

    sql_statement = f"""SELECT
\t{',\n\t'.join(internalNames)}
\t,LTRIM({case_statement_specific},CHAR(10)) AS DEFICIENCY_MININNG_MESSAGE
\t,CASE 
\t\tWHEN {status_conditions} THEN 'check'
\t\tELSE 'ok'
\tEND AS STATUS
FROM
$schema.$table"""

    # create the report
    report_display_name = f"(DEFICIENCIES) GLOBAL"
    report_internal_name = internal_name(report_display_name)

    createOrUpdateReport(
        config=config,
        projectname=projectname,
        displayName=report_display_name,
        internalName=report_internal_name,
        querySyntax=sql_statement,
        description=f"Deficiency Mining Report for  project '{projectname}'",
    )

    createOrUpdateRule(
        config=config,
        projectname=projectname,
        displayName="Global",
        ruleSourceInternalName=report_internal_name,
        ruleGroup="01 Global",
        description=f"Deficiency Mining Rule for project '{projectname}'",
    )

    logging.info(f"Project {projectname}: {len(frags_checked)} checks implemented...")
    return len(frags_checked)


def getNEMOStepsFrompAMigrationStatusFile(file: str) -> list[str]:
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook["Status Datenübernahme"]

    data = []
    for row in worksheet.iter_rows(
        min_row=10, max_row=300, min_col=1, max_col=10, values_only=True
    ):
        data.append(row)

    # Create a DataFrame from the extracted data
    columns = [
        worksheet.cell(row=9, column=i).value for i in range(1, 11)
    ]  # Headers in row 9
    dataframe = pd.DataFrame(data, columns=columns)

    # Drop rows where "Importreihenfolge" is NaN or empty
    if "Importreihenfolge" in dataframe.columns:
        dataframe = dataframe.dropna(subset=["Importreihenfolge"])
    else:
        raise ValueError("The column 'Importreihenfolge' does not exist in the data.")

    if "Name des Importprograms / Name der Erfassungsmaske" in dataframe.columns:
        nemosteps = dataframe[dataframe["Migrationsart"] == "NEMO"][
            "Name des Importprograms / Name der Erfassungsmaske"
        ].to_list()
        replacements = {
            "european article numbers": "global trade item numbers",
            "part-storage areas relationship": "part-storage areas relationships",
        }

        nemosteps = [
            replacements[item] if item in replacements else item for item in nemosteps
        ]

        return nemosteps
    else:
        raise ValueError(
            "The column 'Name des Importprograms / Name der Erfassungsmaske' does not exist in the data."
        )


def updateMappingForMigman(
    config: Config,
    fields: list[str] = None,
):

    # iterate all fields to map and check whether there is already a project
    projectList = getProjectList(config=config)["displayName"].to_list()
    for field in fields:
        projectName = f"Mapping {field}"
        logging.info(f"checking for project '{projectName}'")

        unique_values = []
        
        # if project already exists then preserve data
        if projectName in projectList:
            mappedvalues = LoadReport(config=config,projectname=projectName,report_name="(MAPPING) Non-NULL target values")
            print(df)
            continue
            
        if not projectName in projectList:
            logging.info(f"No project found. Create it.")
            createProject(
                config=config,
                projectname=projectName,
                description=f"Mapping for field '{field}'",
            )

            createImportedColumn(
                config=config,
                projectname=projectName,
                displayName="sourcevalue",
                internalName="sourcevalue",
                importName="sourcevalue",
                dataType="string",
                description="values from source system",
            )

            createImportedColumn(
                config=config,
                projectname=projectName,
                displayName="targetvalue",
                internalName="targetvalue",
                importName="targetvalue",
                dataType="string",
                description="mapped values for target system",
            )

        # let's get some data now
        for project in projectList:
            if project in ["Business Processes", "Master Data"] or project.startswith(
                "Mapping "
            ):
                continue

            imported_columns = getImportedColumns(config=config, projectname=project)[
                "displayName"
            ].to_list()
            result = next(
                (
                    entry
                    for entry in imported_columns
                    if re.match(rf"^{re.escape(field)} \(\d{{3}}\)$", entry)
                ),
                None,
            )
            if result:
                logging.info(f"Found field '{result}' in project '{project}'")

                # create report in that project for that field to get data
                sqlquery = f"""SELECT DISTINCT 
	{internal_name(result)} AS SOURCEVALUE
FROM 
	$schema.PROJECT_{internal_name(project)}
WHERE 
    {internal_name(result)}  IS NOT NULL"""

                displayname = display_name(
                    f"(MAPPING) Distinct Values for mapping of {field}"
                )
                internalname = internal_name(displayname)
                createOrUpdateReport(
                    config=config,
                    projectname=project,
                    displayName=displayname,
                    internalName=internalname,
                    querySyntax=sqlquery,
                    description="load distinct values of field {result} to support mapping of field {field}",
                )

                df = LoadReport(
                    config=config, projectname=project, report_name=displayname
                )

                unique_values.extend(df["SOURCEVALUE"].tolist())

        mappingdf = pd.DataFrame(unique_values, columns=["sourcevalue"])
        mappingdf["targetvalue"] = ""

        # Write to a temporary file and upload
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_file_path = os.path.join(temp_dir, "tempfile.csv")

            mappingdf.to_csv(
                temp_file_path,
                index=False,
                sep=";",
                na_rep="",
            )
            logging.info(
                f"dummy file {temp_file_path} written for project '{projectName}'. Uploading data to NEMO now..."
            )

            ReUploadFile(
                config=config,
                projectname=projectName,
                filename=temp_file_path,
                update_project_settings=False,
            )
            logging.info(f"upload to project {projectName} completed")
            
        # create a report to load all data from this table
        sqlquery = f"""SELECT 
    sourcevalue
    , targetvalue
FROM
    $schema.$table
WHERE
    targetvalue is not NULL
"""
        displayname = "(MAPPING) Non-NULL target values"
        internalname = internal_name(displayname)
        createOrUpdateReport(
            config=config,
            projectname=projectName,
            displayName=displayname,
            internalName=internalname,
            querySyntax=sqlquery,
            description="load all records that have a non-NULL target value",
        )
