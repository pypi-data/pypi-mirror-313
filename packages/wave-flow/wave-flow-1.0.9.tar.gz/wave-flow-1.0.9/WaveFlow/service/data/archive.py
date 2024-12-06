from copy import deepcopy
import os
import json
import time
import openpyxl
import pandas as pd
from typing import Literal

from WaveFlow.service.data.To import To

class Archive:
    
    def __init__(self, file) -> None:
        self.__DesignatedFile = file
        self.__DictWithData:dict[int|str] = dict()
        self.__Dataframe:pd.DataFrame
        self.__Filename:str
        self.__FileType:str
        self.__FileMetaData:dict
        self.__Delimiter:str = ''
        self.__FilesGenerated:list = list()
        self.__background_run()


    def __background_run(self):
        self.__type_file()
        self.__name_of_file()
        self.__metadata_of_file()


    def __name_of_file(self):
        self.__Filename = str(self.__DesignatedFile).replace(self.__FileType, "")[:-1]
    

    def __type_file(self):
        self.__FileType = self.__DesignatedFile.split('.')[-1]
    

    def __metadata_of_file(self):    
        file_size = os.path.getsize(self.__DesignatedFile)
        file_name = os.path.basename(self.__DesignatedFile)
        last_modified = time.ctime(os.path.getmtime(self.__DesignatedFile))
        file_permissions = oct(os.stat(self.__DesignatedFile).st_mode)[-3:]

        metadata = {
            "file_name": file_name,
            "file_size": f"{file_size} bytes",
            "file_type": self.__FileType,
            "last_modified": last_modified,
            "file_permissions": file_permissions}
        
        match self.__FileType:
            case 'csv':
                df = pd.read_csv(self.__DesignatedFile)
                
                metadata.update({
                    "columns": df.columns.tolist()[0].split(';'),
                    "num_rows": df.shape[0],
                    "num_columns": df.shape[1]
                    }
                )

            case 'xlsx':
                wb = openpyxl.load_workbook(self.__DesignatedFile)
                
                metadata.update({
                    "num_sheets": len(wb.sheetnames),
                    "sheet_names": wb.sheetnames,
                    "columns": list(),
                    }
                )
                
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    columns = [cell.value for cell in ws[1]]
                    metadata['columns'].extend(columns)
                    metadata.update({
                        f"{sheet}_dimensions": ws.dimensions,
                        f"{sheet}_max_rows": ws.max_row,
                        f"{sheet}_max_columns": ws.max_column
                    }
                )
                    

            case 'json':
                with open(self.__DesignatedFile, 'r') as file:
                    data = json.load(file)
                
                metadata.update({
                    "num_keys": len(data),
                    "columns": list(data.keys())
                    }
                )
    
            case _:
                pass
            
        self.__FileMetaData = metadata
    

    def __turnDataframeIntoDict(self):
        self.__DictWithData = self.__Dataframe.to_dict(orient='list')
        self.__refactoringDictToPatterns()


    def __wrapWithDelimiter(self, key:str):
        return f'{self.__Delimiter}{key}{self.__Delimiter}'
    

    def __refactoringDictToPatterns(self):
        for k, v in self.__DictWithData.items():
            typeOfValue = type(v[0])
            
            self.__DictWithData.update({k:{'type_column':typeOfValue,
                                           'original_data':v,
                                           'data_handled':deepcopy(v),
                                           'additional_parameters':{'bold':False,
                                                                    'italic':False,
                                                                    'size':0,
                                                                    'font':'',
                                                                    },
                                            }
                                        }
                                    )
        self.__updateKeyWithDelimiter()


    def __updateKeyWithDelimiter(self):
        for k in self.__DictWithData:
            self.__DictWithData[k]['key_w/Delimiter'] = self.__wrapWithDelimiter(k)
        
    
    def transformData(self, keyColumn:str, funcProvided:To):
        """This function can receive an lambda function to be used. e.g.:\n\n
        handler = DataHandler(r'e.g/bd.xlsx')\n
        handler.readFile()\n
        handler.getArchive().transformData(keyColumn="INFO", funcProvided=lambda x: DoSomething(x))
        
        Instead of this, can be used the providade functions in class < To >, check:\n\n
        handler = DataHandler(r'e.g/bd.xlsx')\n
        handler.readFile()\n
        handler.getArchive().transformData("DATE", lambda x: To.Date().to_personalizedFormat(x, '%d de %B de %Y'))"""
        
        for i, obj in enumerate(self.__DictWithData[keyColumn]['original_data']):
            self.__DictWithData[keyColumn]['data_handled'][i] = funcProvided(obj)
        self.__DictWithData[keyColumn]['type_column'] = type(self.__DictWithData[keyColumn]['data_handled'][0])


    def getFileType(self) -> str:
        return self.__FileType


    def getMetaData(self) -> list[str]:
        return self.__FileMetaData


    def getFilename(self) -> str:
        return self.__Filename
    

    def getDesignatedFile(self):
        return self.__DesignatedFile
    

    def getData(self) -> dict:
        if self.__DictWithData:
            return self.__DictWithData
        raise ReferenceError('No Dict Available')
    

    def getDataFrame(self) -> pd.DataFrame:
        if not self.__Dataframe.empty:
            return self.__Dataframe
        raise ReferenceError('No DataFrame Available')
    

    def setDataframe(self, df):
        self.__Dataframe = df
        self.__turnDataframeIntoDict()
    
    
    def getDelimiter(self):
        return self.__Delimiter
    
    
    def getFilesGenerated(self):
        return self.__FilesGenerated    

            
    def setDelimiter(self, newDelimiter:str):
        """This method fills both sides of the keys. \n\n(e.g.: If you pass '==' as newDelimiter, it becomes >>> ==KeyHere==)"""
        self.__Delimiter = str(newDelimiter)
        self.__updateKeyWithDelimiter()


    def setAdditionalParameters(self, keyColumn:str, parameterToChange:Literal["font", "size", "italic", "bold"], newValueToParameter):
        self.__DictWithData[keyColumn]['additional_parameters'][parameterToChange] = newValueToParameter
