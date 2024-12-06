import json,os
from openpyxl import Workbook, load_workbook


xl_file = '/storage/emulated/0/bk_return_sales/bk_return_sales.xlsx'
json_file = "/storage/emulated/0/data.json"

if os.path.exists(json_file):
	with open(json_file,"r") as f:
		data = json.load(f)
		if os.path.exists(xl_file):
			wb = load_workbook(xl_file)
			ws = wb["BEKI"]
		else:
			print("excel file not found")
			
			exit()
else:
	print("json file not found")
	
	exit()
	


v1 = 'knorraio'
v2 = 'knorrchicken'
v3 = 'lbt70g'
v4 = 'lbl70g'
v5 = 'lbt150g'
v6 = 'lbl150g'
v7 = 'lbt20g'
v8 = 'lbl20g'
v9 = 'luxsc70g'
v10 = 'luxst70g'
v11 = 'luxsc150g'
v12 = 'luxst150g'
v13 = 'signal60g'
v14 = 'signal140g'
v15 = 'shacoc350ml'
v16 = 'concoc350ml'
v17 = 'shaavo350ml'
v18 = 'conavo350ml'
v19= 'shacoc700ml'
v20 = 'concoc700ml'
v21 = 'shaavo700ml'
v22 = 'conavo700ml'
v23 = 'shacoc15ml'
v24 = 'concoc15ml'
v25 = 'sunlightbar_200g'
v26 = 'sunlight_40g'
v27 = 'sunlight90g'
v28 = 'sunlight160g'
v29 = 'sunlight500g'
v30 = 'sunlight1kg'
v31 = 'sunlight5kg'
v32 = 'omo40g'
v33 = 'omo100g'
v34 = 'omo500g'
v35 = 'omo1kg'
v36 = 'omo3kg'






def print_pcs():
		        for i in data.keys():
		                if i == v1:
		                	d = data[v1]
		                	ws['L3']=d
		                if i == v2:
		                	d = data[v2]
		                	ws['L4']=d
		                if i == v3:
		                	d = data[v3]
		                	ws['L5']=d
		                if i == v4:
		                	d = data[v4]
		                	ws['L6']=d
		                if i == v5:
		                	d = data[v5]
		                	ws['L7']=d
		                if i == v6:
		                	d = data[v6]
		                	ws['L8']=d
		                if i == v7:
		                	d = data[v7]
		                	ws['L9']=d
		                if i == v8:
		                	d = data[v8]
		                	ws['L10']=d
		                if i == v9:
		                	d = data[v9]
		                	ws['L11']=d
		                if i == v10:
		                	d = data[v10]
		                	ws['L12']=d
		                if i == v11:
		                	d = data[v11]
		                	ws['L13']=d
		                if i == v12:
		                	d = data[v12]
		                	ws['L14']=d
		                if i == v13:
		                	d = data[v13]
		                	ws['L15']=d
		                if i == v14:
		                	d = data[v14]
		                	ws['L16']=d
		                if i == v15:
		                	d = data[v15]
		                	ws['L17']=d
		                if i == v16:
		                	d = data[v16]
		                	ws['L18']=d
		                if i == v17:
		                	d = data[v17]
		                	ws['L19']=d
		                if i == v18:
		                	d = data[v18]
		                	ws['L20']=d
		                if i == v19:
		                	d = data[v19]
		                	ws['L21']=d
		                if i == v20:
		                	d = data[v20]
		                	ws['L22']=d
		                if i == v21:
		                	d = data[v21]
		                	ws['L23']=d
		                if i == v22:
		                	d = data[v22]
		                	ws['L24']=d
		                if i == v23:
		                	d = data[v23]
		                	ws['L25']=d
		                if i == v24:
		                	d = data[v24]
		                	ws['L26']=d
		                if i == v25:
		                	d = data[v25]
		                	ws['L27']=d
		                if i == v26:
		                	d = data[v26]
		                	ws['L28']=d
		                if i == v27:
		                	d = data[v27]
		                	ws['L29']=d
		                if i == v28:
		                	d = data[v28]
		                	ws['L30']=d
		                if i == v29:
		                	d = data[v29]
		                	ws['L31']=d
		                if i == v30:
		                	d = data[v30]
		                	ws['L32']=d
		                if i == v31:
		                	d = data[v31]
		                	ws['L33']=d
		                if i == v32:
		                	d = data[v32]
		                	ws['L34']=d
		                if i == v33:
		                	d = data[v33]
		                	ws['L35']=d
		                if i == v34:
		                	d = data[v34]
		                	ws['L36']=d
		                if i == v35:
		                	d = data[v35]
		                	ws['L37']=d
		                if i == v36:
		                	d = data[v36]
		                	ws['L38']=d
		                                          
	                
	                
	                

if __name__ == "__main__":
	print_pcs()
	wb.save(xl_file)